from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font,colors,PatternFill


excel = load_workbook(filename='words.xlsx')

sheet1 = excel['Sheet1']
sheet2 = excel['Sheet2']


def search_words(word, index):
    try:
        req = requests.get('https://endic.naver.com/search.nhn?sLn=kr&isOnlyViewEE=N&query={}'.format(word))
        html = req.text
        soup = BeautifulSoup(html, 'html.parser')
        meaning = soup.select('.fnt_k05', limit=2)
        playlist = soup.select('.btn_side_play', limit=1)
        pronounciation = soup.select('.fnt_e25', limit=1)
        example = get_example(soup)
        sheet2.cell(row=index, column=3).value = meaning[0].contents[0]
        sheet2.cell(row=index, column=4).value = pronounciation[0].contents[0]
        sheet2.cell(row=index, column=4).hyperlink = playlist[0]['playlist']
        sheet2.cell(row=index, column=4).font = Font(color=colors.BLUE)
        sheet2.cell(row=index, column=5).value = example
        print("ok")
    except:
        sheet2.cell(row=index, column=1).fill = PatternFill(patternType='solid',
                                                            fgColor=colors.RED)
        print("단어를 찾을 수 없습니다.")


def get_example(soup):
    example = soup.select('.fnt_e07', limit=1)
    sentence = ""
    try:
        sentence += example[0].contents[0]
        sentence += example[0].contents[1].contents[0]
        sentence += example[0].contents[2]
        return sentence
    except:
        return sentence


def main():
    for r in sheet1.rows:
        index = r[0].value
        word = r[1].value
        if (r[0].value == None):
            sheet2.cell(row=1, column=2).value = "단어"
            sheet2.cell(row=1, column=3).value = "의미"
            continue
        sheet2.cell(row=index + 1, column=1).value = index
        sheet2.cell(row=index + 1, column=2).value = word
        search_words(word, index + 1)
        excel.save("words.xlsx")


if __name__ == "__main__":
    main()





